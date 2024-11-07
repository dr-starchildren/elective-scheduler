from docx import Document
from docx.shared import Pt, Inches
from ortools.sat.python import cp_model
from itertools import combinations
import numpy as np
import pandas as pd # pip install pandas xlrd
import copy
import openpyxl
import utils
import os
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment

# If you've already generated a schedule, just use the hardcoded solution
# This will save time and resources
run_schedule = False

"""
    Adds a constraint to the schedule model to ensure mutually exclusive electives are scheduled properly.

    This function iterates through each day in the schedule and enforces that, for any set of electives
    defined as mutually exclusive, at most one elective from the set can be scheduled per day.

    Function Details:
    - For each day, the function:
        - Finds the index of each elective in the mutually exclusive list within the electives_data.
        - Skips an elective if it's not found in electives_data.
        - Sums the scheduled values for each elective-counselor combination on the current day.
    - The sum of all selected mutually exclusive electives' schedules for a day is constrained to be at most 1,
      meaning only one of the electives can be scheduled on a given day.
"""
def mutually_exclusive_electives(schedule_model, num_days, counselors_combinations, elective_indexes, mutually_exclusive):
    for day in range(num_days):
        electives_sums = []
        for elective_name in mutually_exclusive:
            # Find the index of the elective in the list of electives
            idx = elective_indexes.get(elective_name, -1)
            if idx == -1:
                continue
            # Sum over all counselor combinations for this activity on this day
            electives_sums.append(sum(scheduled[(day, idx, combo)] for combo in counselors_combinations[idx]))

        # Add a constraint that at most one of these activities can occur on this day
        schedule_model.Add(sum(electives_sums) <= 1)

"""
    Adds a constraint to the schedule model for a specific elective on a specific day.

    This function ensures that at least one counselor combination is assigned to the given elective on the specified day.

    Example:
    >>> schedule_specific_day(model, counselors_combinations, elective_index, 2)
    # Adds a constraint to ensure that "Art" is scheduled on day 2 with at least one counselor combination.

    Function Details:
    - Creates a list of scheduling constraints for each counselor combination available for the elective on the given day.
    - Uses `AddBoolOr` to ensure that at least one of the counselor combinations is scheduled for this elective on the specified day.
"""
def schedule_specific_day(schedule_model, counselors_combinations, elective_index, day):
    day_constraints = [
        scheduled[(day, elective_index, combo)]
        for combo in counselors_combinations[elective_index]
    ]
    schedule_model.AddBoolOr(day_constraints)


# Number of camp days
num_camp_days = 5

electives_data = utils.get_spreadsheet_data(sheet_name='electives')
counselor_data = utils.get_spreadsheet_data(sheet_name='counselors')
camper_data = utils.get_spreadsheet_data(sheet_name='campers')
counselor_rankings = utils.get_rankings('counselor_rankings.csv')
counselor_rankings = [r[1:] for r in counselor_rankings]

elective_indexes = {}
for index, elective in electives_data.iterrows():
    elective_indexes[elective['Name']] = index

required_electives = []
for index, elective in electives_data.iterrows():
    if elective['Min Runnings'] > 0:
        required_electives.append((elective['Name'], elective['Min Runnings']))

counselor_names = counselor_data.apply(lambda c: f"{c['nameFirst']} {c['nameLast']}", axis=1).tolist()
camper_names = camper_data.apply(lambda c: f"{c['Name']}", axis=1).tolist()
all_cabins = camper_data.apply(lambda c: f"{c['Cabin']}", axis=1).tolist()
color_groups = camper_data.apply(lambda c: f"{c['Color']}", axis=1).tolist()

num_electives = electives_data.shape[0]
num_counselors = counselor_data.shape[0]
num_campers = camper_data.shape[0]
# This allows us to pick a schedule that has flexibility in spaces
# Because if we exact match to number of campers, finding an optimal schedule is harder
min_campers_per_day = int(num_campers + num_campers*.2)
max_campers_per_day = int(num_campers + num_campers*.5)


counselors_combinations = {}
# Define the model
schedule_model = cp_model.CpModel()

if run_schedule:
    # Generate all possible combinations of counselors for each elective
    for elective_id, elective_name in electives_data.iterrows():
        counselor_name = elective_name['Counselor']
        num_counselors_required = elective_name['# of Counselors']
        try:
            if counselor_name.lower() == 'any':  # Any counselor can teach it
                if num_counselors_required == 1:
                    counselors_combinations[elective_id] = [(i,) for i in range(num_counselors)]
                else:
                    # Generate combinations for any counselors, meeting the required number
                    counselors_combinations[elective_id] = list(combinations(range(num_counselors), num_counselors_required))
            else:  # Specific counselor is required
                specific_counselor_index = counselor_names.index(counselor_name)
                if num_counselors_required == 1:
                    # Only this specific counselor is needed
                    counselors_combinations[elective_id] = [(specific_counselor_index)]
                else:
                    # Generate combinations that include the specific counselor and meet the total required
                    # Exclude the specific counselor from the range to avoid duplication
                    other_counselors = [i for i in range(num_counselors) if i != specific_counselor_index]
                    # Generate combinations of the remaining counselors needed - 1 (since we already have the specific counselor)
                    other_combinations = list(combinations(other_counselors, num_counselors_required - 1))
                    # Include the specific counselor in each combination
                    counselors_combinations[elective_id] = [(specific_counselor_index,) + combo for combo in other_combinations]
        except ValueError:
            print(f"Error: Counselor '{counselor_name}' not found in the list of counselors.")

    # Create all possible options for activity day (d), elective (e), and counselor combo (c)
    scheduled = {}
    for day in range(num_camp_days):
        for elective_id in range(num_electives):
            for combo in counselors_combinations[elective_id]:
                scheduled[(day, elective_id, combo)] = schedule_model.NewBoolVar(f'scheduled_d{day}_e{elective_id}_c{"_".join(map(str, combo))}')


    """
        Constraint #1 
        Each activity does not exceed its max occurrences
    """
    for elective_id in range(num_electives):
        max_occurrences = electives_data.loc[elective_id, 'Max Runnings']
        # print(f"Max occurrences for elective {e}: {max_occurrences}") 
        schedule_model.Add(sum(scheduled[(d, elective_id, combo)] for d in range(num_camp_days) for combo in counselors_combinations[elective_id]) <= max_occurrences)
    
    """
        Constraint #2
        No counselor double-booking
    """
    for counselor_id in range(num_counselors):
        for day in range(num_camp_days):
            schedule_model.Add(sum(scheduled[(day, elective_id, combo)] for elective_id in range(num_electives) for combo in counselors_combinations[elective_id] if counselor_id in combo) <= 1)

    """
        Constraint #3
        We must have enough spaces for all students must be in an activity
    """
    for day in range(num_camp_days):
        schedule_model.Add(sum(scheduled[(day, elective_id, combo)] * electives_data.loc[elective_id, 'Max Campers'] for elective_id in range(num_electives) for combo in counselors_combinations[elective_id]) >= min_campers_per_day)

    """
        Constraint #3
        These activities can't run on the same day
    """
    ropes_electives = ['Rock Wall', 'Zip Line', 'Giant Swing']
    mutually_exclusive_electives(schedule_model, num_camp_days, counselors_combinations, elective_indexes, ropes_electives)
    sports_electives = ['Gaga Ball', 'Sponge Ball']
    mutually_exclusive_electives(schedule_model, num_camp_days, counselors_combinations, elective_indexes, sports_electives)

    """
        Constraint #3
        Each activity can only occur once per day
    """
    for day in range(num_camp_days):
        for elective_id in range(num_electives):
            schedule_model.Add(sum(scheduled[(day, elective_id, combo)] for combo in counselors_combinations[elective_id]) <= 1)

    """
        Constraint #4
        Make sure required electives are scheduled
    """
    for elective_name, total in required_electives:
        elective_index = elective_indexes[elective_name]
        if elective_index == -1:
            continue
        # Sum all scheduling variables for this elective across all days and combinations
        elective_scheduled = sum(scheduled[(day, elective_index, combo)]
                                for day in range(num_camp_days)
                                for combo in counselors_combinations[elective_index])
        # Add constraint to ensure this elective is scheduled at least once
        schedule_model.Add(elective_scheduled >= total)

    """
        Constraint #5
        Minimize ranking for counselors
    """
    preference_cost = sum(scheduled[(day, elective_index, counselor_combo)] * counselor_rankings[counselor][elective_index]
        for day in range(num_camp_days)
        for elective_index,_ in electives_data.iterrows()
        for counselor_combo in counselors_combinations[elective_index]
        for counselor in counselor_combo
        if (day, elective_index, counselor_combo) in scheduled)
    schedule_model.Minimize(preference_cost)

    """
        Constraint #6
        Counselor can't teach same class more than two times, this will mix up the schedule
        This is based on counselor rankings. If you counselors didn't rank things, this is not necessary
    """
    for elective_id, _ in electives_data.iterrows():
        for counselor in range(num_counselors): 
            teaching_instances = []
            for day in range(num_camp_days):
                for combo in counselors_combinations[elective_id]:
                    if counselor in combo:
                        var = scheduled.get((day, elective_id, combo))
                        if var is not None:
                            teaching_instances.append(var)      
            if teaching_instances:
                schedule_model.Add(sum(teaching_instances) <= 2)

    """
        Constraint #7
        schedule activities for a specific days
    """
    for index, elective in electives_data.iterrows():
        if elective['Days to Run'].lower() != 'any':
            required_days = elective['Days to Run'].split(',')
            for required_day in required_days:
                schedule_specific_day(schedule_model, counselors_combinations, elective_indexes[elective['Name']], utils.day_of_week(required_day.strip()))
    
    # Solve the model
    solver = cp_model.CpSolver()
    # Take up to 10 minutes to find solution
    solver.parameters.max_time_in_seconds = 600.0
    status = solver.Solve(schedule_model)

    elective_schedule = {day: [] for day in range(num_camp_days)} 
    full_schedule = []

    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        for day in range(num_camp_days):
            total_spaces = 0
            print(f'Day {day + 1}:')
            for elective_id, elective_name in electives_data.iterrows():
                for combo in counselors_combinations[elective_id]:
                    if solver.Value(scheduled[(day, elective_id, combo)]):
                        elective_schedule[day].append(elective_id)
                        full_schedule.append((day, elective_id, combo))
                        try:
                            for counselor_id in combo:
                                if counselor_rankings[counselor_id][elective_id] > 5:
                                    print('\tFailed Top 5:',counselor_names[counselor_id], elective_name['Name'], counselor_rankings[counselor_id][elective_id])
                                
                            counselors_str = ', '.join(counselor_names[c] for c in combo)
                        except:
                            print('err',counselor_names[counselor_id])
                        print(f'  Elective: { elective_name["Name"] }, Counselors: {counselors_str}')
                        total_spaces = total_spaces + elective_name['Max Campers']

            print('Total Spaces:', total_spaces)
            print()
            print("Elective Schedule: Raw Data")
            print("if you don't want to run the schedule creator every time, save these values to elective_schedule and full_schedule")
            print('elective_schedule =', elective_schedule)
            print('full_scheule =', full_schedule)
            print()
    else:
        print("No solution found.")

else:
    elective_schedule = {0: [0, 3, 9, 11, 12, 13, 15, 20], 1: [1, 7, 9, 11, 12, 13, 18, 19], 2: [0, 4, 7, 8, 11, 12, 15, 17], 3: [0, 1, 5, 7, 10, 12, 14, 21], 4: [2, 9, 11, 12, 13, 14, 16, 18]}
    full_schedule = [(0, 0, (2, 15)), (0, 3, (5, 8, 14)), (0, 9, (6, 7)), (0, 11, (11, 16)), (0, 12, (12, 13)), (0, 13, (1, 17)), (0, 15, (0, 18)), (0, 20, (3, 4)), (1, 1, (1, 4)), (1, 7, (0, 2)), (1, 9, (15, 16)), (1, 11, (5, 10)), (1, 12, (14, 18)), (1, 13, (9, 11)), (1, 18, (7, 12, 17)), (1, 19, (8, 3)), (2, 0, (7, 18)), (2, 4, (8, 5)), (2, 7, (3, 17)), (2, 8, (0, 4, 10)), (2, 11, (9, 11)), (2, 12, (14, 15)), (2, 15, (1, 13)), (2, 17, (6, 16)), (3, 0, (1, 6)), (3, 1, (2, 7)), (3, 5, (8, 10)), (3, 7, (4, 13)), (3, 10, (15, 14)), (3, 12, (12, 18)), (3, 14, (0, 5, 9)), (3, 21, (11, 17)), (4, 2, (15, 16)), (4, 9, (3, 14)), (4, 11, (1, 6)), (4, 12, (2, 13)), (4, 13, (7, 12)), (4, 14, (0, 10, 18)), (4, 16, (8, 4)), (4, 18, (5, 9, 17))]


# After we have a schedule, we can try to optimize camper rankings
rankings = utils.get_rankings('rankings.csv')
rankings = [r[1:] for r in rankings]

# Create a new model for ranking campers
ranking_model = cp_model.CpModel()

# Create all possible combinations of camper, day, and elective running on a specific day
assignments = {}
for camper_id in range(num_campers):
    for day in range(num_camp_days):
        for elective_name in range(num_electives):
            if elective_name in elective_schedule[day]:
                var_name = f'camper_{camper_id}_day_{day}_elective_{elective_name}'
                assignments[(camper_id, day, elective_name)] = ranking_model.NewBoolVar(var_name)

"""
    Constraint #1 
    No duplicate activities
    A camper cannot attend the same activity twice

"""
for camper_id in range(num_campers):
    for elective_name in range(num_electives):
        # Sum all the assignments of this elective for this camper across all days
        total_assignments = sum(assignments[(camper_id, day, elective_name)]
                                for day in range(num_camp_days)
                                if (camper_id, day, elective_name) in assignments)
        # Add a constraint that this total must be no more than 1
        ranking_model.Add(total_assignments <= 1)

# in case a ranking wasn't included, we'll just assign a max rank
max_rank = 100  
camper_rankings = [
    [int(rank) if not np.isnan(rank) else max_rank for rank in camper] 
    for camper in rankings
]

"""
    Constraint #2
    Number of campers in an activity must be between min and max allowed
    
"""
for day in range(num_camp_days):
    for elective_name in range(num_electives):
        # Get maximum capacity for the current elective      
        max_capacity = electives_data.loc[elective_name, 'Max Campers']
        min_capacity = electives_data.loc[elective_name, 'Min Campers']

        # Sum up all assignments for this elective on this day
        total_assigned = sum(assignments[(camper_id, day, elective_name)]
                             for camper_id in range(num_campers)
                             if (camper_id, day, elective_name) in assignments)
        # Add constraint that this total does not exceed the maximum capacity
        ranking_model.Add(total_assigned <= max_capacity)
        # Make sure we hit minimum capicity of classes
        # There's a quirk where the total_assigned can be '0', which is a string.
        # We have to throw these out otherwise the model can never resolve because all classes have min_capacity > 0
        if not isinstance(total_assigned, int):
            ranking_model.Add(total_assigned >= min_capacity)

"""
    Constraint #3
    Each student is assigned to exactly one activity that is available according to the elective_schedule
    
"""
for day in range(num_camp_days):
    for counselor_id in range(num_campers):
        ranking_model.Add(sum(assignments[(counselor_id, day, elective_id)] for elective_id in elective_schedule[day] if (counselor_id, day, elective_id) in assignments) == 1)


"""
    Constraint #4 
    This is where the magic happens - Minimize Camper Ranking Preferences
    We try to optimize for the lowest total overall camper ranking
    This does not weight a difference between campers or prioritize if a camper got their #1 rank.
    It just tries to optimize for overall lowest ranks across the whole camp
    There can be quirks like 4 people getting their #1 rank is worse than one person getting their #3 rank item, but
    in practice those four people would then have their higher ranks of other electives picked, which would increase overall cost
    It's not the perfectly balanced solution, but works pretty well out of the box without other considerations
"""
total_preference = sum(assignments[(camper_id, day, elective_id)] * camper_rankings[camper_id][elective_id]
    for day in range(num_camp_days)
    for camper_id in range(num_campers)
    for elective_id in elective_schedule[day]
    if (camper_id, day, elective_id) in assignments)
ranking_model.Minimize(total_preference)

solver = cp_model.CpSolver()
# Stop looking for solutions after 10 minutes
solver.parameters.max_time_in_seconds = 600.0
# Run the solver
status = solver.Solve(ranking_model)

# check if we have a solution
if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
    # Create a new Document
    doc = Document()
    doc.sections.top_margin = Inches(0.5)

    camper_schedule = {}
    week_schedule = {}
    counselor_schedule = {}

    # Iterate over each day
    for day in range(num_camp_days):
        print(f"Day {day + 1}:")
        electives_today = {}
        week_schedule[utils.day_of_week(day)] = []
        # Collect which campers are assigned to each elective on this day
        for elective_name in range(num_electives):
            campers_in_elective = []
            for camper_id in range(num_campers):
                # Check if the assignment variable is True
                if (camper_id, day, elective_name) in assignments and solver.Value(assignments[(camper_id, day, elective_name)]):
                    campers_in_elective.append(camper_id)

                    if camper_id in camper_schedule:
                        camper_schedule[camper_id] = camper_schedule[camper_id] + [elective_name]
                    else:
                        camper_schedule[camper_id] = [elective_name]
            
            # If any campers are assigned, save it under the current elective
            if campers_in_elective:
                elective_name = electives_data.loc[elective_name, 'Name']
                electives_today[elective_name] = campers_in_elective
        
        # Print the activities and campers for this day
        for elective_name, camper_id in electives_today.items():
            week_schedule[utils.day_of_week(day)].append(elective_name)
            sign_in_sheet = []
            camper_list = [camper_names[c] for c in camper_id]
            for counselor_id in camper_id:
                sign_in_sheet.append(camper_names[counselor_id])
            
            elective_data = electives_data.loc[electives_data['Name'] == elective_name]

            elective_counselors = []
            if not elective_data.empty:
                temp_counselors = [ entry[2] for entry in full_schedule if entry[0] == day and entry[1] == elective_data.index[0] ]
                for counselor_id in temp_counselors[0]:
                    elective_counselors.append(counselor_names[counselor_id].split(' ')[0])
                    if counselor_id not in counselor_schedule:
                        counselor_schedule[counselor_id] = {day:''}
                    counselor_schedule[counselor_id][day] = elective_name
            
            utils.word_doc_output(doc, elective_name, sign_in_sheet, day, elective_counselors)

            print(f"  {elective_name}: enrolled ({str(len(camper_id))}), max ({elective_data['Max Campers'].iloc[0]}) , min ({elective_data['Min Campers'].iloc[0]}):")
            print(f"      {', '.join(camper_list)}")
        print("") 
    # Save the document
    doc.save('Daily_Signin_Sheets.docx')

    #print(week_schedule)
    ws = pd.DataFrame.from_dict(week_schedule, orient='index')

    with pd.ExcelWriter(f'Week_Schedule.xlsx', engine='openpyxl') as writer:
        transpose = ws.transpose()
        transpose.to_excel(writer, sheet_name='Schedule')

    # Output the schedule for each individual camper
    not_first_choice = []
    not_second_choice = []
    excel_schedules = []
    excel_df_base = {
        'Camper':[],
        'Cabin':[],
        'Color': [],
        'Mon': [],
        'Tues': [],
        'Wed': [],
        'Thur': [],
        'Fri':[]
    }
    excel_df = copy.deepcopy(excel_df_base)
    excel_columns = 5

    cabins = set(camper_data['Cabin'].tolist())
    cabin_schedules_df = {}
    for cabin in cabins:
        cabin_schedules_df[cabin] = {}

    for camper_id,_ in camper_data.iterrows():
        first_choice = False
        second_choice = False
        #print(f"{all_campers[camper_id].strip()}")
        #print(f"Camper Preferences {camper_rankings[camper_id]}:")
        excel_df['Camper'] += ['Camper:'] + [utils.printable_camper(camper_id, camper_names)]
        excel_df['Cabin'] += ['Cabin:'] + [all_cabins[camper_id]]
        excel_df['Color'] += ['Color:'] + [color_groups[camper_id]]
        cabin_schedules_df[all_cabins[camper_id]][utils.printable_camper(camper_id, camper_names)] = []
        for day, elective_name in enumerate(camper_schedule[camper_id]):
            day_string = utils.day_of_week(day)
            excel_df[day_string] += [day_string+':'] + [electives_data.loc[elective_name, 'Name']]
            cabin_schedules_df[all_cabins[camper_id]][utils.printable_camper(camper_id, camper_names)].append(electives_data.loc[elective_name, 'Name'])
            if camper_rankings[camper_id][elective_name] == 1:
                first_choice = True
            if camper_rankings[camper_id][elective_name] == 2:
                second_choice = True
            
            #print(f" Day {day+1}: {electives[elective][0]} ({camper_rankings[camper_id][elective]})")
        #print('\n')
        if not first_choice:
            not_first_choice.append(camper_names[camper_id])
        if not second_choice:
            not_second_choice.append(camper_names[camper_id])

        if camper_id % excel_columns == excel_columns-1 or camper_id == len(camper_names)-1:
            df = pd.DataFrame(excel_df)
            df_transposed = df.transpose()
            df_transposed.columns = [f'Camper: {col}' for col in df_transposed.columns]
            # Open the existing workbook and append the new data
            # Check if the file exists, and create it if not
            output_file = 'Campers_Schedule.xlsx'
            if not os.path.exists(output_file):
                # Create an empty DataFrame and save it to create the file
                df = pd.DataFrame()
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False)
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:   
                # Check if the 'Campers Schedule' sheet already exists
                if 'Campers Schedule' in writer.book.sheetnames:
                    startrow = writer.book['Campers Schedule'].max_row
                else:
                    startrow = 0
                df_transposed.loc[len(df)] = None
                # Append data without overwriting; increment index to avoid overwriting headers if sheet exists
                df_transposed.to_excel(writer, sheet_name='Campers Schedule', index=False, startrow=startrow, header=False if startrow > 0 else True)
                
            excel_df = copy.deepcopy(excel_df_base)

    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    wb = openpyxl.load_workbook('Campers_Schedule.xlsx', )
    sheet = wb['Campers Schedule'] 
    sheet.delete_rows(1)
    if 'Sheet1' in wb.sheetnames:
        del wb['Sheet1']
    # Dimensions of sections and the area to cover
    row_increment = len(excel_df) + 1
    col_increment = 2
    total_rows = 125  # Total rows to cover, adjust as needed
    total_cols = 10  # Total columns to cover, adjust as needed

    # Apply the border to each section
    for start_row in range(1, total_rows + 1, row_increment):
        for start_col in range(1, total_cols + 1, col_increment):
            # Define the end of the section
            end_row = start_row + row_increment - 1
            end_col = start_col + col_increment - 1

            # Loop through cells in defined section
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    # Assigning borders to cells at the edge of the section
                    cell = sheet.cell(row, col)
                    if row == start_row:  # Top edge
                        cell.border = cell.border + Border(top=Side(style='thin'))
                        cell.font = Font(bold=True)
                    if row == end_row:  # Bottom edge
                        cell.border = cell.border + Border(bottom=Side(style='thin'))
                    if col == start_col:  # Left edge
                        cell.border = cell.border + Border(left=Side(style='thin'))
                    if col == end_col:  # Right edge
                        cell.border = cell.border + Border(right=Side(style='thin'))


    # Format the columns to fit the content
    utils.resize_columns(sheet)
    # Save the workbook
    wb.save('Campers_Schedule.xlsx')


    #########################
    # Counselor schedules
    #########################
    excel_df_base = {
        'Counselor':[],
        'Mon': [],
        'Tues': [],
        'Wed': [],
        'Thur': [],
        'Fri':[]
    }
    excel_df = copy.deepcopy(excel_df_base)
    excel_columns = 5
    print(counselor_schedule)
    for counselor_id in sorted(counselor_schedule.keys()):
        excel_df['Counselor'] += ['Counselor:'] + [counselor_names[counselor_id].split(' ')[0]]
        for day in range(num_camp_days):
            if day in counselor_schedule[counselor_id].keys():
                excel_df[utils.day_of_week(day)] += [utils.day_of_week(day)+':'] + [counselor_schedule[counselor_id][day]]
            else:
                excel_df[utils.day_of_week(day)] += [utils.day_of_week(day)+':'] + ['None']

        if counselor_id % excel_columns == excel_columns-1 or counselor_id == num_counselors-1:
            df = pd.DataFrame(excel_df)
            df_transposed = df.transpose()
            df_transposed.columns = [f'Counselor: {col}' for col in df_transposed.columns]
            # Open the existing workbook and append the new data
            output_file = 'Counselors_Schedule.xlsx'
            if not os.path.exists(output_file):
                # Create an empty DataFrame and save it to create the file
                df = pd.DataFrame()
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False)
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # Check if the 'Schedule' sheet already exists
                if 'Schedule' in writer.book.sheetnames:
                    startrow = writer.book['Schedule'].max_row
                else:
                    startrow = 0
                df_transposed.loc[len(df)] = None
                # Append data without overwriting; increment index to avoid overwriting headers if sheet exists
                df_transposed.to_excel(writer, sheet_name='Schedule', index=False, startrow=startrow, header=False if startrow > 0 else True)
                
            excel_df = copy.deepcopy(excel_df_base)

    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    wb = openpyxl.load_workbook('Counselors_Schedule.xlsx', )
    sheet = wb['Schedule']
    sheet.delete_rows(1)
    if 'Sheet1' in wb.sheetnames:
        del wb['Sheet1']
    
    # Dimensions of sections and the area to cover
    row_increment = len(excel_df) + 1
    col_increment = 2
    total_rows = 29  # Total rows to cover, adjust as needed
    total_cols = 10  # Total columns to cover, adjust as needed

    # Apply the border to each section
    for start_row in range(1, total_rows + 1, row_increment):
        for start_col in range(1, total_cols + 1, col_increment):
            # Define the end of the section
            end_row = start_row + row_increment - 1
            end_col = start_col + col_increment - 1

            # Loop through cells in defined section
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    # Assigning borders to cells at the edge of the section
                    cell = sheet.cell(row, col)
                    if row == start_row:  # Top edge
                        cell.border = cell.border + Border(top=Side(style='thin'))
                        cell.font = Font(bold=True)
                    if row == end_row:  # Bottom edge'w'
                        cell.border = cell.border + Border(bottom=Side(style='thin'))
                    if col == start_col:  # Left edge
                        cell.border = cell.border + Border(left=Side(style='thin'))
                    if col == end_col:  # Right edge
                        cell.border = cell.border + Border(right=Side(style='thin'))

    # Format the columns to fit the content
    utils.resize_columns(sheet)
    # Save the workbook
    wb.save('Counselors_Schedule.xlsx')

    # Create the schedule for each cabin
    for cabin in cabin_schedules_df.keys():
        with pd.ExcelWriter(f'{cabin}_Schedule.xlsx', engine='openpyxl') as writer:
            sdf = pd.DataFrame(cabin_schedules_df[cabin])
            sdf_transposed = sdf.transpose()
            sdf_transposed.columns = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
            sdf_transposed.to_excel(writer, sheet_name='Schedule')

    colors = ['FFFFFF','ADD8E6']
    for cabin in cabin_schedules_df.keys():
        wb = openpyxl.load_workbook(f'{cabin}_Schedule.xlsx')
        desired_font = Font(size=15)
        for sheet in wb.sheetnames:
            worksheet = wb[sheet]
            i = 0
            for row in worksheet.iter_rows():
                
                fill = PatternFill(start_color=colors[i%len(colors)], end_color=colors[i%len(colors)], fill_type='solid')
                for cell in row:
                    cell.font = desired_font
                    cell.fill = fill
                    worksheet.row_dimensions[cell.row].height = 20
                    cell.alignment = Alignment(vertical='center', horizontal='left', wrapText=False)
                i = i + 1
        
        wb.save(f'{cabin}_Schedule.xlsx')

    print("Didn't get first choice")
    print(', '.join(not_first_choice))
    print("Didn't get second choice")
    for counselor_id in not_first_choice:
        if counselor_id in not_second_choice:
            print(counselor_id)
    print()


else:
    print("No solution found.")
